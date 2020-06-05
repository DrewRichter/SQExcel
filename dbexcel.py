import sqlite3
import openpyxl
from openpyxl import load_workbook
import re
import datetime
from flask import Flask, render_template, url_for, request, redirect
from flask_sqlalchemy import SQLAlchemy
from sqlalchemy import create_engine

 # three slashes for relative path. 4 for absolute

class DBExcel:
    def __init__(self, wb_name, file_name, ):

        self.wb_name=wb_name
        wb=openpyxl.load_workbook(wb_name)
        self.wb=wb



        self.file_name=file_name
        conn = sqlite3.connect(file_name)
        self.conn=conn
        print("")
        c = conn.cursor()

        self.c=c

    def query(self, selected_tables, from_tables, conditions=None, sheet_name=None):
        command= "SELECT"
        for item in selected_tables:
            command += " "+ item
            if not item==selected_tables[-1]:
                command+= ","
        command+=" FROM"
        for item in from_tables:
            command+= " "+item
            if not item==from_tables[-1]:
                command+= ","
        if not conditions==None:
            command+= " WHERE"
            for item in conditions:
                command += " " + item
                if not item == conditions[-1]:
                    command += " AND"

        statement=self.c.execute(command)
        if sheet_name==None:
            sheet=self.wb.create_sheet()
        #new_sheet.append([(string[:string.index(".")] for string in selected_tables if "." in string)])
        else:
            sheet = self.wb[sheet_name]
        sheet.append(selected_tables)
        return_statement=[]
        for row in statement:
            sheet.append(list(row))
            return_statement.append(list(row))
        self.wb.save(self.wb_name)
        self.conn.commit()
        return return_statement



    def execute(self, statement):
        self.c.execute(statement)
        self.conn.commit()


    def insert_values(self, table_name, values):
        pass
        statement= "INSERT INTO {table} VALUES(".format(table=table_name)
        s=statement
        for value in values:
            s+=value
            if not value == values[-1]:
                s+=", "
        s+=")"
        print(s)
        self.c.execute(s)
        self.conn.commit()



    def load_table(self, table_name,sheet_name):
        sheet=self.wb[sheet_name]
        sheet.append(self.c.execute("PRAGMA table_info(?)",table_name).fetchall())
        for item in self.c.execute("SELECT * FROM ?", table_name).fetchall():
            sheet.append(item)
        self.wb.save()
        self.c.commit()


    def create_table(self, table_name, values):
        statement = "CREATE TABLE {table}("
        for value in values:
            statement += " " + str(value)
            if not value == values[-1]:
                statement += ","
        statement = statement.format(table=table_name) + ")"
        self.c.execute(statement)